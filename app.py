from __future__ import annotations

import os
import sqlite3
from datetime import datetime
from functools import wraps
from io import BytesIO

from flask import (
    Flask,
    flash,
    g,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DATABASE = os.path.join(BASE_DIR, "database.db")

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "cambia-questa-chiave")
app.config["ADMIN_USERNAME"] = os.environ.get("ADMIN_USERNAME", "admin")
app.config["ADMIN_PASSWORD"] = os.environ.get("ADMIN_PASSWORD", "admin123")

RIONI = [
    "POZZOLUNGO NORD",
    "POZZOLUNGO SUD",
    "IANA",
    "CHIANCA",
    "ZITA ROSA",
    "CENTRO",
    "CONSOLAZIONE",
    "PATULA",
    "CUPA QUARTARARU",
]

SPORTS = {
    "Calcio": {"fee": 10.0, "is_double": False},
    "Padel": {"fee": 40.0, "is_double": True},
    "Burraco": {"fee": 5.0, "is_double": True},
    "Tiro con l'arco": {"fee": 10.0, "is_double": False},
    "Biliardino": {"fee": 20.0, "is_double": True},
    "Scopa": {"fee": 5.0, "is_double": True},
    "Volley": {"fee": 10.0, "is_double": False},
    "Scacchi": {"fee": 10.0, "is_double": False},
    "1vs1": {"fee": 5.0, "is_double": False},
    "Tennis": {"fee": 10.0, "is_double": False},
    "Ludopoli": {"fee": 5.0, "is_double": False},
}

BELONGING_OPTIONS = [
    "Residente o domiciliato attualmente",
    "Residente o domiciliato per lungo periodo e recentemente trasferito",
    "Inserito stabilmente nel tessuto sociale (non residente o non domiciliato)",
]

SHIRT_SIZES = ["XS", "S", "M", "L", "XL", "XXL"]

MAP_URL = "https://www.google.com/maps/@40.2878732,17.9905045,15z/data=!3m1!4b1!4m2!6m1!1s1AiNo30X3Qloy24nwtxUo06kHozB_yC4?authuser=1&entry=ttu&g_ep=EgoyMDI1MDUwNi4wIKXMDSoASAFQAw%3D%3D"
SHIRT_PRICE = 5.0


@app.template_filter("currency")
def currency_filter(value: float) -> str:
    try:
        return f"€ {float(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return str(value)


def get_db() -> sqlite3.Connection:
    if "db" not in g:
        g.db = sqlite3.connect(DATABASE)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(error=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db() -> None:
    db = sqlite3.connect(DATABASE)
    db.executescript(
        """
        CREATE TABLE IF NOT EXISTS registrations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at TEXT NOT NULL,
            email TEXT NOT NULL,
            sport TEXT NOT NULL,
            rione TEXT NOT NULL,
            base_fee REAL NOT NULL,
            shirt_price REAL NOT NULL,
            total_fee REAL NOT NULL,
            player1_name TEXT NOT NULL,
            player1_cf TEXT NOT NULL,
            player1_phone TEXT NOT NULL,
            player1_belonging TEXT NOT NULL,
            player1_address TEXT NOT NULL,
            player1_shirt INTEGER NOT NULL DEFAULT 0,
            player1_shirt_size TEXT,
            player2_name TEXT,
            player2_cf TEXT,
            player2_phone TEXT,
            player2_belonging TEXT,
            player2_address TEXT,
            player2_shirt INTEGER NOT NULL DEFAULT 0,
            player2_shirt_size TEXT,
            confirm_fee INTEGER NOT NULL DEFAULT 1,
            confirm_rione_check INTEGER NOT NULL DEFAULT 1,
            confirm_approval INTEGER NOT NULL DEFAULT 1,
            privacy_ok INTEGER NOT NULL DEFAULT 1,
            images_ok INTEGER NOT NULL DEFAULT 1,
            liability_ok INTEGER NOT NULL DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS app_settings (
            key TEXT PRIMARY KEY,
            value TEXT
        );
        """
    )
    db.commit()
    db.close()


def ensure_db_schema() -> None:
    db = get_db()
    columns = {row["name"] for row in db.execute("PRAGMA table_info(registrations)").fetchall()}

    extra_columns = {
        "player1_shirt_size": "TEXT",
        "player2_shirt_size": "TEXT",
    }

    for col, col_type in extra_columns.items():
        if col not in columns:
            db.execute(f"ALTER TABLE registrations ADD COLUMN {col} {col_type}")

    db.execute(
        """
        CREATE TABLE IF NOT EXISTS app_settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
        """
    )
    db.commit()


def get_setting(key: str, default: str = "") -> str:
    db = get_db()
    ensure_db_schema()
    row = db.execute("SELECT value FROM app_settings WHERE key = ?", (key,)).fetchone()
    return row["value"] if row else default


def set_setting(key: str, value: str) -> None:
    db = get_db()
    ensure_db_schema()
    db.execute(
        "INSERT INTO app_settings(key, value) VALUES(?, ?) ON CONFLICT(key) DO UPDATE SET value = excluded.value",
        (key, value),
    )
    db.commit()


def get_registration_open_status() -> tuple[bool, str]:
    manual_closed = get_setting("registrations_closed", "0") == "1"
    auto_close_at = get_setting("auto_close_at", "").strip()

    if manual_closed:
        return False, "Le iscrizioni sono state chiuse dagli organizzatori."

    if auto_close_at:
        try:
            close_dt = datetime.fromisoformat(auto_close_at)
            if datetime.now() >= close_dt:
                return False, f"Le iscrizioni sono chiuse dal {close_dt.strftime('%d/%m/%Y alle %H:%M')}."
        except ValueError:
            pass

    return True, ""


def login_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if not session.get("admin_logged_in"):
            return redirect(url_for("admin_login"))
        return view_func(*args, **kwargs)

    return wrapped


@app.route("/")
def index():
    is_open, closed_message = get_registration_open_status()
    return render_template(
        "index.html",
        sports=SPORTS,
        rioni=RIONI,
        belonging_options=BELONGING_OPTIONS,
        shirt_price=SHIRT_PRICE,
        shirt_sizes=SHIRT_SIZES,
        map_url=MAP_URL,
        registrations_open=is_open,
        closed_message=closed_message,
    )


@app.post("/submit")
def submit_registration():
    is_open, closed_message = get_registration_open_status()
    if not is_open:
        flash(closed_message or "Le iscrizioni sono chiuse.", "error")
        return redirect(url_for("index"))

    form = request.form
    sport_name = form.get("sport", "")
    sport = SPORTS.get(sport_name)
    if not sport:
        flash("Seleziona uno sport valido.", "error")
        return redirect(url_for("index"))

    if form.get("confirm_fee") != "Si":
        flash("Per procedere devi confermare la quota di iscrizione.", "error")
        return redirect(url_for("index"))

    required_yes_fields = {
        "confirm_rione_check": "Conferma verifica appartenenza al rione",
        "confirm_approval": "Conferma approvazione organizzatori",
        "privacy_ok": "Autorizzazione privacy",
        "images_ok": "Autorizzazione immagini",
        "liability_ok": "Conferma responsabilità",
    }
    for key, label in required_yes_fields.items():
        if form.get(key) != "Si":
            flash(f"Devi accettare: {label}.", "error")
            return redirect(url_for("index"))

    player1_name = form.get("player1_name", "").strip()
    player1_cf = form.get("player1_cf", "").strip().upper()
    player1_phone = form.get("player1_phone", "").strip()
    player1_belonging = form.get("player1_belonging", "").strip()
    player1_address = form.get("player1_address", "").strip()
    email = form.get("email", "").strip()
    rione = form.get("rione", "").strip()
    wants_shirt_1 = 1 if form.get("shirt_player1") == "Si" else 0
    player1_shirt_size = form.get("player1_shirt_size", "").strip()

    if not all([player1_name, player1_cf, player1_phone, player1_belonging, player1_address, email, rione]):
        flash("Compila tutti i campi obbligatori del 1° giocatore.", "error")
        return redirect(url_for("index"))

    if wants_shirt_1 and player1_shirt_size not in SHIRT_SIZES:
        flash("Se il 1° giocatore vuole la maglia devi selezionare una taglia valida.", "error")
        return redirect(url_for("index"))
    if not wants_shirt_1:
        player1_shirt_size = ""

    player2_name = form.get("player2_name", "").strip()
    player2_cf = form.get("player2_cf", "").strip().upper()
    player2_phone = form.get("player2_phone", "").strip()
    player2_belonging = form.get("player2_belonging", "").strip()
    player2_address = form.get("player2_address", "").strip()
    wants_shirt_2 = 1 if form.get("shirt_player2") == "Si" else 0
    player2_shirt_size = form.get("player2_shirt_size", "").strip()

    if sport["is_double"]:
        if not all([player2_name, player2_cf, player2_phone, player2_belonging, player2_address]):
            flash("Per questo sport di coppia devi compilare anche i dati del 2° giocatore.", "error")
            return redirect(url_for("index"))
        if wants_shirt_2 and player2_shirt_size not in SHIRT_SIZES:
            flash("Se il 2° giocatore vuole la maglia devi selezionare una taglia valida.", "error")
            return redirect(url_for("index"))
    else:
        player2_name = player2_cf = player2_phone = player2_belonging = player2_address = ""
        wants_shirt_2 = 0
        player2_shirt_size = ""

    if not wants_shirt_2:
        player2_shirt_size = ""

    base_fee = sport["fee"]
    total_fee = base_fee + (wants_shirt_1 * SHIRT_PRICE) + (wants_shirt_2 * SHIRT_PRICE)

    db = get_db()
    ensure_db_schema()
    db.execute(
        """
        INSERT INTO registrations (
            created_at, email, sport, rione, base_fee, shirt_price, total_fee,
            player1_name, player1_cf, player1_phone, player1_belonging, player1_address, player1_shirt, player1_shirt_size,
            player2_name, player2_cf, player2_phone, player2_belonging, player2_address, player2_shirt, player2_shirt_size,
            confirm_fee, confirm_rione_check, confirm_approval, privacy_ok, images_ok, liability_ok
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            email,
            sport_name,
            rione,
            base_fee,
            SHIRT_PRICE,
            total_fee,
            player1_name,
            player1_cf,
            player1_phone,
            player1_belonging,
            player1_address,
            wants_shirt_1,
            player1_shirt_size,
            player2_name,
            player2_cf,
            player2_phone,
            player2_belonging,
            player2_address,
            wants_shirt_2,
            player2_shirt_size,
            1,
            1,
            1,
            1,
            1,
            1,
        ),
    )
    db.commit()

    return render_template(
        "success.html",
        sport_name=sport_name,
        rione=rione,
        total_fee=total_fee,
        email=email,
    )


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "")
        if username == app.config["ADMIN_USERNAME"] and password == app.config["ADMIN_PASSWORD"]:
            session["admin_logged_in"] = True
            return redirect(url_for("admin_dashboard"))
        flash("Credenziali non valide.", "error")
    return render_template("admin_login.html")


@app.route("/admin/logout")
def admin_logout():
    session.clear()
    return redirect(url_for("admin_login"))


@app.route("/admin")
@login_required
def admin_dashboard():
    db = get_db()
    ensure_db_schema()

    selected_sport = request.args.get("sport", "")
    if selected_sport and selected_sport in SPORTS:
        rows = db.execute(
            "SELECT * FROM registrations WHERE sport = ? ORDER BY created_at DESC",
            (selected_sport,),
        ).fetchall()
    else:
        rows = db.execute("SELECT * FROM registrations ORDER BY created_at DESC").fetchall()

    summary = db.execute(
        """
        SELECT sport,
               COUNT(*) as registrations_count,
               SUM(total_fee) as total_collected,
               SUM(player1_shirt + player2_shirt) as shirts_count
        FROM registrations
        GROUP BY sport
        ORDER BY sport ASC
        """
    ).fetchall()

    rioni_shirts = []
    for rione in RIONI:
        counts = dict(db.execute(
            """
            SELECT COALESCE(player1_shirt_size, '') as size, COUNT(*) as qty
            FROM registrations
            WHERE rione = ? AND player1_shirt = 1
            GROUP BY COALESCE(player1_shirt_size, '')
            """,
            (rione,)
        ).fetchall())
        counts2 = dict(db.execute(
            """
            SELECT COALESCE(player2_shirt_size, '') as size, COUNT(*) as qty
            FROM registrations
            WHERE rione = ? AND player2_shirt = 1
            GROUP BY COALESCE(player2_shirt_size, '')
            """,
            (rione,)
        ).fetchall())
        merged = {size: counts.get(size, 0) + counts2.get(size, 0) for size in SHIRT_SIZES}
        total = sum(merged.values())
        rioni_shirts.append({"rione": rione, "sizes": merged, "total": total})

    is_open, closed_message = get_registration_open_status()
    auto_close_at = get_setting("auto_close_at", "")

    return render_template(
        "admin_dashboard.html",
        rows=rows,
        sports=SPORTS,
        selected_sport=selected_sport,
        summary=summary,
        rioni_shirts=rioni_shirts,
        shirt_sizes=SHIRT_SIZES,
        registrations_open=is_open,
        closed_message=closed_message,
        auto_close_at=auto_close_at,
    )


@app.post("/admin/toggle-registrations")
@login_required
def toggle_registrations():
    current = get_setting("registrations_closed", "0")
    set_setting("registrations_closed", "0" if current == "1" else "1")
    flash("Stato iscrizioni aggiornato.", "success")
    return redirect(url_for("admin_dashboard"))


@app.post("/admin/set-auto-close")
@login_required
def set_auto_close():
    auto_close_at = request.form.get("auto_close_at", "").strip()
    if auto_close_at:
        try:
            dt = datetime.fromisoformat(auto_close_at)
            set_setting("auto_close_at", dt.isoformat(timespec="minutes"))
            flash("Data di chiusura automatica salvata.", "success")
        except ValueError:
            flash("Formato data non valido.", "error")
    else:
        set_setting("auto_close_at", "")
        flash("Data di chiusura automatica rimossa.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/edit/<int:registration_id>", methods=["GET", "POST"])
@login_required
def edit_registration(registration_id: int):
    db = get_db()
    ensure_db_schema()
    row = db.execute("SELECT * FROM registrations WHERE id = ?", (registration_id,)).fetchone()
    if not row:
        flash("Iscrizione non trovata.", "error")
        return redirect(url_for("admin_dashboard"))

    if request.method == "POST":
        sport_name = request.form.get("sport", "")
        sport = SPORTS.get(sport_name)
        if not sport:
            flash("Sport non valido.", "error")
            return redirect(url_for("edit_registration", registration_id=registration_id))

        wants_shirt_1 = 1 if request.form.get("player1_shirt") == "Si" else 0
        wants_shirt_2 = 1 if request.form.get("player2_shirt") == "Si" else 0
        player1_shirt_size = request.form.get("player1_shirt_size", "").strip() if wants_shirt_1 else ""
        player2_shirt_size = request.form.get("player2_shirt_size", "").strip() if wants_shirt_2 else ""

        player2_name = request.form.get("player2_name", "").strip()
        player2_cf = request.form.get("player2_cf", "").strip().upper()
        player2_phone = request.form.get("player2_phone", "").strip()
        player2_belonging = request.form.get("player2_belonging", "").strip()
        player2_address = request.form.get("player2_address", "").strip()

        if not sport["is_double"]:
            player2_name = player2_cf = player2_phone = player2_belonging = player2_address = ""
            wants_shirt_2 = 0
            player2_shirt_size = ""

        total_fee = sport["fee"] + (wants_shirt_1 * SHIRT_PRICE) + (wants_shirt_2 * SHIRT_PRICE)

        db.execute(
            """
            UPDATE registrations SET
                email = ?, sport = ?, rione = ?, base_fee = ?, total_fee = ?,
                player1_name = ?, player1_cf = ?, player1_phone = ?, player1_belonging = ?, player1_address = ?, player1_shirt = ?, player1_shirt_size = ?,
                player2_name = ?, player2_cf = ?, player2_phone = ?, player2_belonging = ?, player2_address = ?, player2_shirt = ?, player2_shirt_size = ?
            WHERE id = ?
            """,
            (
                request.form.get("email", "").strip(),
                sport_name,
                request.form.get("rione", "").strip(),
                sport["fee"],
                total_fee,
                request.form.get("player1_name", "").strip(),
                request.form.get("player1_cf", "").strip().upper(),
                request.form.get("player1_phone", "").strip(),
                request.form.get("player1_belonging", "").strip(),
                request.form.get("player1_address", "").strip(),
                wants_shirt_1,
                player1_shirt_size,
                player2_name,
                player2_cf,
                player2_phone,
                player2_belonging,
                player2_address,
                wants_shirt_2,
                player2_shirt_size,
                registration_id,
            ),
        )
        db.commit()
        flash("Iscrizione modificata correttamente.", "success")
        return redirect(url_for("admin_dashboard"))

    return render_template(
        "edit_registration.html",
        row=row,
        sports=SPORTS,
        rioni=RIONI,
        belonging_options=BELONGING_OPTIONS,
        shirt_sizes=SHIRT_SIZES,
    )


@app.post("/admin/delete/<int:registration_id>")
@login_required
def delete_registration(registration_id: int):
    db = get_db()
    db.execute("DELETE FROM registrations WHERE id = ?", (registration_id,))
    db.commit()
    flash("Iscrizione eliminata.", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/export")
@login_required
def export_excel():
    db = get_db()
    ensure_db_schema()
    all_rows = db.execute("SELECT * FROM registrations ORDER BY sport, created_at DESC").fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Riepilogo"

    header_fill = PatternFill("solid", fgColor="111111")
    header_font = Font(color="FFFFFF", bold=True)

    riepilogo_headers = ["Sport", "Iscrizioni", "Totale incassi", "Maglie richieste"]
    ws.append(riepilogo_headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    sport_groups = {sport: [] for sport in SPORTS.keys()}
    for row in all_rows:
        sport_groups[row["sport"]].append(row)

    for sport, rows in sport_groups.items():
        total = sum(r["total_fee"] for r in rows)
        shirts = sum((r["player1_shirt"] or 0) + (r["player2_shirt"] or 0) for r in rows)
        ws.append([sport, len(rows), total, shirts])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"iscrizioni_torneo_rioni_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/export-anagrafica")
@login_required
def export_players_excel():
    db = get_db()
    ensure_db_schema()
    all_rows = db.execute("SELECT * FROM registrations ORDER BY sport, created_at DESC").fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Giocatori"

    header_fill = PatternFill("solid", fgColor="111111")
    header_font = Font(color="FFFFFF", bold=True)

    headers = [
        "Data iscrizione",
        "Sport",
        "Rione",
        "Ruolo",
        "Nome e cognome",
        "Codice fiscale",
        "Email",
        "Telefono",
        "Maglia richiesta",
        "Taglia maglia",
    ]

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for r in all_rows:
        ws.append([
            r["created_at"],
            r["sport"],
            r["rione"],
            "1° giocatore",
            r["player1_name"],
            r["player1_cf"],
            r["email"],
            r["player1_phone"],
            "Si" if r["player1_shirt"] else "No",
            r["player1_shirt_size"] or "",
        ])
        if r["player2_name"]:
            ws.append([
                r["created_at"],
                r["sport"],
                r["rione"],
                "2° giocatore",
                r["player2_name"],
                r["player2_cf"],
                r["email"],
                r["player2_phone"],
                "Si" if r["player2_shirt"] else "No",
                r["player2_shirt_size"] or "",
            ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"anagrafica_giocatori_torneo_rioni_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/export-magliette-rioni")
@login_required
def export_rioni_shirts_excel():
    db = get_db()
    ensure_db_schema()

    wb = Workbook()
    ws = wb.active
    ws.title = "Magliette per rione"

    header_fill = PatternFill("solid", fgColor="111111")
    header_font = Font(color="FFFFFF", bold=True)

    headers = ["Rione", *SHIRT_SIZES, "Totale magliette"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for rione in RIONI:
        size_totals = {size: 0 for size in SHIRT_SIZES}
        p1 = db.execute(
            "SELECT player1_shirt_size as size, COUNT(*) as qty FROM registrations WHERE rione = ? AND player1_shirt = 1 GROUP BY player1_shirt_size",
            (rione,),
        ).fetchall()
        p2 = db.execute(
            "SELECT player2_shirt_size as size, COUNT(*) as qty FROM registrations WHERE rione = ? AND player2_shirt = 1 GROUP BY player2_shirt_size",
            (rione,),
        ).fetchall()

        for row in p1:
            if row["size"] in size_totals:
                size_totals[row["size"]] += row["qty"]
        for row in p2:
            if row["size"] in size_totals:
                size_totals[row["size"]] += row["qty"]

        total = sum(size_totals.values())
        ws.append([rione, *[size_totals[size] for size in SHIRT_SIZES], total])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"magliette_rione_per_rione_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/health")
def health():
    return {"status": "ok"}


if __name__ == "__main__":
    if not os.path.exists(DATABASE):
        init_db()
    else:
        with app.app_context():
            ensure_db_schema()

    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port, debug=False)
